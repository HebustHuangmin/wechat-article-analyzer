#!/usr/bin/env python
# -*- coding: UTF-8 -*-
'''
@Project ：crawl_OA 
@File    ：main.py
@IDE     ：PyCharm 
@Author  ：李思璇
@Date    ：2025/3/4 10:26 
'''
import os
import openpyxl
import pandas as pd
import datetime
import json
from selenium import webdriver
from lxml import html
from selenium.webdriver.common.by import By
import xlsxwriter as xw
import requests
import re
import time
with open('./account.txt','r') as fp:
    content = fp.readline()
#     获取用户名
account = content.split("  ")[0]
# 获取用户密码
passwd = content.split("  ")[1]

class C_ookie:
    # 初始化
    def __init__(self):
        self.html = ''
    # 获取cookie
    def get_cookie(self):
        url = 'https://mp.weixin.qq.com'
        Browner = webdriver.Chrome()
        Browner.get(url)
        Browner.maximize_window()
        Browner.find_element(By.XPATH, "//a[text()='使用账号登录']").click()
        # 获取账号输入框
        ID = Browner.find_element(By.NAME, "account")
        # ID = Browner.find_element_by_name('account')
        # 获取密码输入框
        PW = Browner.find_element(By.NAME, "password")
        # 输入账号
        id = account
        pw = passwd
        ID.send_keys(id)
        PW.send_keys(pw)
        # 获取登录button，点击登录
        Browner.find_element(By.CLASS_NAME,'btn_login').click()
        # Browner.find_element_by_class_name('btn_login').click()
        # 一定要扫描二维码，等待扫二维码
        time.sleep(30)
        ck = Browner.get_cookies()

        ck1 = json.dumps(ck)
        with open('ck.txt','w') as f :
            f.write(ck1)
            f.close()
        self.html = Browner.page_source

	# 获取token，在页面中提取
    def Token(self):
        etree = html.etree
        h = etree.HTML(self.html)
        url = h.xpath('//a[@title="首页"]/@href')[0]
        token = re.findall('\d+',url)[0]
        return token
        # with open('token.txt', 'w') as f:
        #     f.write(token[0])
        #     f.close()


C = C_ookie()
C.get_cookie()
# 获取唯一身份id
token = C.Token()

headers = {
    'accept': '*/*',
    'accept-language': 'zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7,en-GB;q=0.6',
    # 'cookie': 'pgv_pvid=3750802949; RK=BNVAjJ6WtZ; ptcz=833eb995f13f94575bd63e4224f6352c0d1298be609a853458c7d4c063ab2150; qq_domain_video_guid_verify=9d41e40d32d0b889; _qimei_uuid42=18a110b031810079763f4b9d70b89eb09914bf7cf6; _qimei_fingerprint=e1a97f2dda1c8109eea730b02ed03c7a; _qimei_h38=19c52253763f4b9d70b89eb002000001e18a11; _qimei_q32=1f5a2715f2dcb13b273025e1753e50ba; _qimei_q36=2146240cff30501c9fde564a30001dd18903; ua_id=20Y8Zc8oyK8ySS2FAAAAAKfulg8idpuBvdrg9YczpH8=; wxuin=36148104706972; mm_lang=zh_CN; noticeLoginFlag=1; _clck=1e5tua8|1|fse|0; uuid=b53ad704247e58307a10b38e93915c10; bizuin=3957993985; ticket=eff3d7a7d3d679d914d1525cbd3597705a01f7ff; ticket_id=gh_234e1231c92b; slave_bizuin=3957993985; cert=3p8eK3jCjf8rmhc3oP1CkOK50FGb_Hyv; remember_acct=2763474254%40qq.com; rand_info=CAESIES2i9FcXNVaRTjHc48k3c2ITGqmk3PBu9vd9AFO2Ucq; data_bizuin=3957993985; data_ticket=VsSfXzo0LvcGoTL820ZAjGPhnogRfrVghtNUzuE7iFgnsrS8LS9NYbQKtF28QyjA; slave_sid=c01iSHg0cExKQzdCc2FRQTdCVkRlOFRlSklveHNNenlrMUZoaEp1ZEd0ODIzSUhnWUVnS2VoMlRfMEJRZjdJcDdZWFBMS1BYT2s2ZjNWMUdtRG9Pa3N6b2Vlc0VfMTdJWEJ6RDhPWWNkWGFfZXIwRGZUSHptUDF1TGZ6bTY4MGxnWlZRNm1xUGhrNE1FSllB; slave_user=gh_234e1231c92b; xid=429c6f819f840d324648bf7e48941a05; openid2ticket_oATLq6_gM3n99BETsC702MzEukXc=THzA9u0lVFJkFjNzzWbrpwzVnOxhPU0xPTbeMEgg5J8=; _clsk=nyzkq5|1736323668874|2|1|mp.weixin.qq.com/weheat-agent/payload/record',
    'priority': 'u=1, i',
    'referer': 'https://mp.weixin.qq.com/cgi-bin/appmsg?t=media/appmsg_edit_v2&action=edit&isNew=1&type=77&createType=8&token=320710992&lang=zh_CN&timestamp=1736323668409',
    'sec-ch-ua': '"Microsoft Edge";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'empty',
    'sec-fetch-mode': 'cors',
    'sec-fetch-site': 'same-origin',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0',
    'x-requested-with': 'XMLHttpRequest',
}

params = {
    'sub': 'list',
    'search_field': 'null',
    'begin': '0',
    'count': '30',
    'query': '',
    'fakeid': 'MzI0MjMzNjczMw==',
    'type': '101_1',
    'free_publish_type': '1',
    'sub_action': 'list_ex',
    # 将身份id作为参数传进去
    'token':str(token),
    'lang': 'zh_CN',
    'f': 'json',
    'ajax': '1',
}
# 71
with open('ck.txt', 'r') as f:
    cookie = f.read()
    f.close()
cookie1 = json.loads(cookie)
cookie_dit = {}
for dic in cookie1:
    key = dic['name']
    value = dic['value']
    cookie_dit[key]=value

if os.path.exists('./高新发布.xlsx'):
    # 加载已有的工作簿
    workbook = openpyxl.load_workbook('./高新发布.xlsx')
    # 获取指定工作表
    worksheet1 = workbook['sheet1']
else:
    # 创建新的工作簿
    workbook = openpyxl.Workbook()
    # 获取活动工作表
    worksheet1 = workbook.active
    # 设置工作表标题
    worksheet1.title = 'sheet1'
    # 设置表头
    title = [ '发布时间', '标题', '备注（链接）']
    for col_num, value in enumerate(title, 1):
        worksheet1.cell(row=1, column=col_num, value=value)

# 读取现有数据（如果文件已存在）
try:
    # 读取 Excel 文件
    df = pd.read_excel('./高新发布.xlsx')
    existing_links = set(df['备注（链接）'].tolist())  # 获取已有链接
except FileNotFoundError:
    print("文件不存在，将创建新文件。")
    existing_links = set()

# 开始请求
response = requests.get('https://mp.weixin.qq.com/cgi-bin/appmsgpublish', params=params, cookies=cookie_dit, headers=headers)
publish_list = json.loads(response.json()['publish_page'])['publish_list']

# 处理数据
total_data = []
k = len(existing_links) + 1  # 序号从已有数据的数量 +1 开始
for i in publish_list:
    article_list = json.loads(i['publish_info'])['appmsgex']
    for j in article_list:
        link = j['link']
        if link not in existing_links:  # 去重
            data = []
            print(j['title'], j['link'], datetime.datetime.fromtimestamp(j['create_time']).strftime("%Y.%m.%d"),k)
            # data.append(str(k))
            data.append(datetime.datetime.fromtimestamp(j['create_time']).strftime("%Y.%m.%d"))
            data.append(j['title'])
            data.append(link)
            total_data.append(data)
            # existing_links.add(link)  # 将新链接加入已有链接集合
            # k += 1
total_data.sort(key=lambda x: datetime.datetime.strptime(x[0], "%Y.%m.%d"))
# 追加数据到 Excel
max_row = worksheet1.max_row
for insertData in total_data:
    max_row += 1
    for col_num, value in enumerate(insertData, 1):
        worksheet1.cell(row=max_row, column=col_num, value=value)

# 保存工作簿
workbook.save('./高新发布.xlsx')
print("操作完成！！！")

# 读取现有数据
# try:
#     # 读取 Excel 文件
#     df = pd.read_excel('./高新发布.xlsx')
#     existing_links = set(df['备注（链接）'].tolist())  # 获取已有链接
#     print(f"已加载现有数据，共有 {len(existing_links)} 条记录")
# except FileNotFoundError:
#     print("文件不存在，将创建新文件。")
#     df = pd.DataFrame(columns=['日期', '标题', '备注（链接）'])  # 创建空DataFrame
#     existing_links = set()
#
# # 开始请求
# response = requests.get('https://mp.weixin.qq.com/cgi-bin/appmsgpublish', params=params, cookies=cookie_dit,
#                         headers=headers)
# publish_list = json.loads(response.json()['publish_page'])['publish_list']
#
# # 处理新数据
# total_data = []
# new_links = set()  # 记录本次新增的链接
#
# for i in publish_list:
#     article_list = json.loads(i['publish_info'])['appmsgex']
#     for j in article_list:
#         link = j['link']
#         if link not in existing_links and link not in new_links:
#             data = [
#                 datetime.datetime.fromtimestamp(j['create_time']).strftime("%Y.%m.%d"),
#                 j['title'],
#                 link
#             ]
#             total_data.append(data)
#             new_links.add(link)  # 记录新增链接
#
# # 检查是否有新数据
# if not total_data:
#     print("没有新数据需要添加")
# else:
#     # 创建新数据的DataFrame
#     new_df = pd.DataFrame(total_data, columns=['日期', '标题', '备注（链接）'])
#
#     # 转换日期列为datetime类型，以便正确排序
#     df['日期'] = pd.to_datetime(df['日期'], format='%Y.%m.%d')
#     new_df['日期'] = pd.to_datetime(new_df['日期'], format='%Y.%m.%d')
#
#     # 合并新旧数据
#     combined_df = pd.concat([df, new_df], ignore_index=True)
#
#     # 按日期升序排序（旧数据在前，新数据在后）
#     combined_df = combined_df.sort_values(by='日期', ascending=True)
#
#     # 将日期列转回字符串格式（保持与原代码一致）
#     combined_df['日期'] = combined_df['日期'].dt.strftime('%Y.%m.%d')
#
#     # 保存到Excel（覆盖原文件）
#     combined_df.to_excel('./高新发布.xlsx', index=False)
#     print(f"成功将 {len(new_df)} 条新数据按时间顺序插入到文件中")